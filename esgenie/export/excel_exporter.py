"""
excel_exporter.py
Template-driven ESG Excel export. Reads kpi_template_v8_FINAL.csv to
map KPIs to the correct tab. Returns bytes for Streamlit download.
"""

import io
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_KPI_TEMPLATE_PATH = Path(__file__).parent.parent / "config" / "kpi_template_v8_FINAL.csv"

# Header style
_HDR_FILL = PatternFill("solid", fgColor="1F4E2C")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)

# Row confidence styles
_FILL_HIGH = PatternFill("solid", fgColor="E2EFDA")
_FILL_MED  = PatternFill("solid", fgColor="FFFF99")
_FILL_LOW  = PatternFill("solid", fgColor="FFE0E0")
_FILL_NONE = PatternFill("solid", fgColor="F2F2F2")

_FONT_NOT_FOUND = Font(italic=True, color="999999")

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")

# dashboard_section keywords → tab name
_TAB_MAP = {
    "Environment": ["Climate", "Biodiversity", "Environmental", "Water"],
    "Social":      ["Social"],
    "Governance":  ["Governance"],
}

_DATA_COLUMNS = [
    "KPI ID", "KPI Canonical Name", "Theme",
    "Value", "Unit", "Year", "Source Page", "Confidence",
]


def _tab_for_section(section: str) -> str:
    """Map a dashboard_section string to the correct tab name."""
    for tab, keywords in _TAB_MAP.items():
        if any(kw.lower() in section.lower() for kw in keywords):
            return tab
    return "Environment"  # fallback


def _row_fill(confidence: str) -> PatternFill:
    c = str(confidence or "").lower()
    if c == "high":
        return _FILL_HIGH
    if c == "medium":
        return _FILL_MED
    if c in ("low", "not_found", ""):
        return _FILL_LOW
    return _FILL_NONE


def _write_header(ws, columns: list[str]) -> None:
    for i, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _WRAP
        cell.border = _BORDER
    ws.row_dimensions[1].height = 22


def _write_data_row(ws, row: int, values: list, confidence: str) -> None:
    fill = _row_fill(confidence)
    not_found = str(confidence or "").lower() in ("not_found", "")
    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=val if val is not None else "")
        cell.fill = fill
        cell.border = _BORDER
        cell.alignment = _WRAP
        if not_found:
            cell.font = _FONT_NOT_FOUND


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value or "")) for c in col_cells), default=10
        )
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max_len + 4, 55)


def export_to_excel(
    extractions: list[dict],
    company: str,
    year: int,
    scores: dict | None = None,
) -> bytes:
    """
    Build a template-driven Excel workbook from ESG extractions.

    Args:
        extractions: List of extraction dicts. Accepted key variants:
                     kpi_id, kpi_canonical_name / metric, kpi_theme / theme,
                     value / value_raw, unit, reporting_year / year,
                     source_page, confidence, dashboard_section (optional).
        company:     Company name for headers.
        year:        Reporting year.
        scores:      Optional output from esg_scorer.score_company().

    Returns:
        Raw bytes of the .xlsx file.
    """
    kpi_df = pd.read_csv(_KPI_TEMPLATE_PATH).set_index("kpi_id")

    # Build lookup: kpi_id -> dashboard_section
    section_map: dict[str, str] = kpi_df["dashboard_section"].to_dict()

    # Normalise extraction dicts
    def _norm(e: dict) -> dict:
        return {
            "kpi_id":     e.get("kpi_id", ""),
            "canonical":  e.get("kpi_canonical_name") or e.get("metric", ""),
            "theme":      e.get("kpi_theme") or e.get("theme", ""),
            "value":      e.get("value") or e.get("value_raw") or "",
            "unit":       e.get("unit", ""),
            "year":       e.get("reporting_year") or e.get("year", year),
            "page":       e.get("source_page", ""),
            "confidence": e.get("confidence", "not_found"),
            "section":    e.get("dashboard_section")
                          or section_map.get(e.get("kpi_id", ""), ""),
        }

    normed = [_norm(e) for e in extractions]

    # Group by tab
    tabs: dict[str, list[dict]] = {"Environment": [], "Social": [], "Governance": []}
    for e in normed:
        tab = _tab_for_section(e["section"])
        tabs[tab].append(e)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Data tabs ---
    for tab_name in ["Environment", "Social", "Governance"]:
        ws = wb.create_sheet(title=tab_name)
        _write_header(ws, _DATA_COLUMNS)

        for row_idx, e in enumerate(tabs[tab_name], start=2):
            val = str(e["value"]) if e["value"] not in (None, "") else "Not Found"
            _write_data_row(
                ws,
                row_idx,
                [
                    e["kpi_id"], e["canonical"], e["theme"],
                    val, e["unit"], e["year"],
                    e["page"], e["confidence"],
                ],
                e["confidence"],
            )

        _auto_width(ws)
        ws.freeze_panes = "A2"

    # --- ESG Score tab ---
    ws_score = wb.create_sheet(title="ESG Score")

    ws_score.merge_cells("A1:D1")
    title_cell = ws_score["A1"]
    title_cell.value = f"{company} — ESG Score Summary ({year})"
    title_cell.font = Font(bold=True, size=14, color="1F4E2C")

    ws_score["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_score["A2"].font = Font(italic=True, color="666666")

    if scores:
        overall = scores.get("overall_score", 0)
        rating, interp = scores.get("overall_rating", "N/A"), scores.get("overall_interpretation", "")

        ws_score["A4"] = "Overall ESG Score"
        ws_score["A4"].font = Font(bold=True, size=12)
        ws_score["B4"] = overall
        ws_score["B4"].font = Font(bold=True, size=20, color="1F4E2C")
        ws_score["C4"] = rating
        ws_score["C4"].font = Font(bold=True, size=16, color="1F4E2C")
        ws_score["D4"] = interp

        ws_score["A6"] = "Theme"
        ws_score["B6"] = "Score"
        ws_score["C6"] = "Rating"
        ws_score["D6"] = "Completeness"
        ws_score["E6"] = "Avg Confidence"
        ws_score["F6"] = "Controversy Adj."
        for col in "ABCDEF":
            cell = ws_score[f"{col}6"]
            cell.font = _HDR_FONT
            cell.fill = _HDR_FILL
            cell.border = _BORDER

        theme_scores = scores.get("theme_scores", {})
        for r_idx, (theme, tdata) in enumerate(theme_scores.items(), start=7):
            row_vals = [
                theme.capitalize(),
                tdata.get("score", 0),
                tdata.get("rating", ""),
                f"{tdata.get('completeness', 0):.0%}",
                f"{tdata.get('avg_confidence', 0):.0%}",
                tdata.get("controversy_deduction", 0),
            ]
            fill = _row_fill(tdata.get("rating", ""))
            for c_idx, val in enumerate(row_vals, 1):
                cell = ws_score.cell(row=r_idx, column=c_idx, value=val)
                cell.border = _BORDER
                cell.alignment = _WRAP

        cont = scores.get("controversy_total_deduction", 0)
        ws_score.cell(row=len(theme_scores) + 8, column=1,
                      value=f"Total Controversy Deduction: {cont}")

    _auto_width(ws_score)

    # Reorder: Score tab first
    wb.move_sheet("ESG Score", offset=-len(wb.sheetnames) + 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
